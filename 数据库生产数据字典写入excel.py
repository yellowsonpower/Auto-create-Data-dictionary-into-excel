import  pandas as  pd
import openpyxl

#**********************读取ddl 写数据字典*********#
ddl="""
CREATE TABLE [dbo].[PsiMab_pass1_pivot](
	[ExcelName] [nvarchar](500) NULL,
	[SheetName] [nvarchar](500) NULL,
	[MarsY] [nvarchar](500) NULL,
	[MarsP] [nvarchar](500) NULL,
	[Version] [nvarchar](500) NULL,
	[Identification] [nvarchar](500) NULL,
	[Category] [nvarchar](500) NULL,
	[Type] [nvarchar](500) NULL,
	[BOH] [nvarchar](500) NULL,
	[P1] [nvarchar](4000) NULL,
	[P2] [nvarchar](4000) NULL,
	[P3] [nvarchar](4000) NULL,
	[P4] [nvarchar](4000) NULL,
	[P5] [nvarchar](4000) NULL,
	[P6] [nvarchar](4000) NULL,
	[P7] [nvarchar](4000) NULL,
	[P8] [nvarchar](4000) NULL,
	[P9] [nvarchar](4000) NULL,
	[P10] [nvarchar](4000) NULL,
	[P11] [nvarchar](4000) NULL,
	[P12] [nvarchar](4000) NULL,
	[P13] [nvarchar](4000) NULL
) ON [PRIMARY]
GO
"""

#print(ddl)


other=ddl[:ddl.index(") ON [PRIMARY]")].replace('CREATE TABLE','')
createtable=ddl[:ddl.index("(")]
tablename=ddl[:ddl.index("(")].replace('CREATE TABLE','')

other=ddl[:ddl.index(") ON [PRIMARY]")].replace(createtable+'(','')

#print('table name:',tablename)


#print(other)


#*****************************读取数据库获取表结构************#

import openpyxl
import os
import sqlalchemy
import pyodbc
import pandas as pd
import pymssql


#*************************链接数据库信息************************#
server   = 
database = 
username = 
password = 

#********************链接数据库******#
cnxn = pymssql.connect(server, username, password, database)
cursor = cnxn.cursor()

#************获取 数据库table表结构 *********#
df = pd.read_sql("""
SELECT *,
COLUMNPROPERTY(OBJECT_ID(TABLE_SCHEMA+'.'+TABLE_NAME), COLUMN_NAME, 'IsPrimaryKey') AS IsPrimaryKey
FROM INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_NAME like 'DMR%'   
or  TABLE_NAME like 'PsiMab%'
"""
, cnxn)

print(df.info())

#**************关闭数据库链接******************#
cnxn.close()




dfneed=df[['TABLE_NAME','COLUMN_NAME','ORDINAL_POSITION','IsPrimaryKey','IS_NULLABLE','DATA_TYPE','CHARACTER_MAXIMUM_LENGTH']]

#print(dfneed)

sheetname1=dfneed[['TABLE_NAME']].drop_duplicates()

sheetname2=sheetname1['TABLE_NAME'].tolist()

#print(dfneed)
#dfn = dfneed[dfneed['TABLE_NAME' == 'PsiMab_test']]
#print(dfn)




import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook



# 使用 ExcelWriter() 方法创建一个新的 ExcelWriter 对象
filepath = 'C:/Users/xunyi/Desktop/DMR数据字典.xlsx'
writer = pd.ExcelWriter(filepath)

dfneed.to_excel(writer, sheet_name="all_table", index=False)
writer._save()

# 遍历 sheetname2 列表中的每个元素
for i in sheetname2:
    # 创建新的工作表，名称为从字符串中获取的前 30 个字符
    #worksheet = workbook.create_sheet(i[:30])
    # 获取 dfneed 数据帧中 TABLE_NAME 列为当前工作表名称的行
    dfn = dfneed[dfneed['TABLE_NAME'] == i]
    dfn.to_excel(writer, sheet_name=i[:30], index=False)
    #print(dfn)

# # 保存 ExcelWriter 对象并关闭它
writer._save()


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('C:/Users/xunyi/Desktop/DMR数据字典.xlsx')
sheet_names = workbook.sheetnames

#print(workbook.worksheets)


for worksheet in workbook.worksheets:
    #worksheet = workbook.active
    #worksheet.title = i
    col=['B','C','D','E','G']
    for nn in col:
        worksheet.column_dimensions[nn].width = 15
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['F'].width = 20

workbook.save('C:/Users/xunyi/Desktop/DMR数据字典.xlsx')
